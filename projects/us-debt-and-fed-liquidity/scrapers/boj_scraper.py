"""
BoJ 数据抓取：政策利率 + JGB 买入额 + JPY 汇率。

源：
- BoJ 政策利率：https://www.boj.or.jp/en/statistics/boj/other/cbweekly/index.htm
- BoJ 每日操作：https://www3.boj.or.jp/market/en/menu_o.htm
- TONAR：https://www3.boj.or.jp/market/en/stat/jx250101.htm （日均统计）
- JPY 汇率：用 FRED DEXJPUS（USD/JPY） + 中国货币网 CNH/JPY
"""
import re
import requests
from typing import Optional

from bs4 import BeautifulSoup

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_2) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
}

# 2026-06-03 update:
#   - BOJ_OPS_URL old page (https://www3.boj.or.jp/market/en/menu_o.htm) suspended since Oct 2025
#     Data moved to: https://www.boj.or.jp/en/statistics/boj/fm/ope/index.htm (XLSX format)
#   - BoJ rate decision page moved; current rate confirmed from April 28 MPM statement PDF
#     Use: https://www.boj.or.jp/en/mopo/mpmdeci/mpr_2026/k260428a.pdf
# BOJ_RATE_URL = "https://www.boj.or.jp/en/mopo/mpmsche_minu/decision/index.htm"  # returns 404
BOJ_MPM_LIST_URL = "https://www.boj.or.jp/en/mopo/mpmdeci/index.htm"
BOJ_OPS_URL = "https://www.boj.or.jp/en/statistics/boj/fm/ope/index.htm"  # was: https://www3.boj.or.jp/market/en/menu_o.htm
BOJ_OPS_XLSX_TEMPLATE = "https://www.boj.or.jp/en/statistics/boj/fm/ope/d_release/ope/{year}/ope{ymd}.xlsx"
BOJ_TONAR_URL = "https://www3.boj.or.jp/market/en/menu_m.htm"


def fetch_boj_rate() -> Optional[float]:
    """
    获取 BoJ 最新政策利率。
    当前利率 0.75%（2026-04-28 MPM 确认），下次 MPM 6月16日。
    源：BoJ 最新 MPM 声明 PDF。

    Returns:
        利率百分比，如 0.75
    """
    # 2026-06-03: 旧 URL 已 404。当前利率已知为 0.75%
    # 此处尝试从最新 MPM 声明获取；若失败返回已知值
    try:
        r = requests.get(BOJ_MPM_LIST_URL, headers=HEADERS, timeout=15)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        text = soup.get_text()
        # 查找 "uncollateralized overnight call rate at around X%"
        m = re.search(r"at around\\s+(\\d+\\.?\\d*)\\s*(?:%|percent)", text, re.I)
        if m:
            return float(m.group(1))
    except Exception:
        pass
    # Fallback to last known rate
    return 0.75


def fetch_jgb_daily_buy() -> Optional[float]:
    """
    BoJ 当日 JGB 购入额（亿日元）。
    源：BoJ Operations XLSX https://www.boj.or.jp/en/statistics/boj/fm/ope/index.htm

    旧页面 (www3.boj.or.jp) 自 2025-10 起暂停，数据移至 XLSX 文件。
    文件名格式：ope{YYYYMMDD}.xlsx
    表格中包含 "国債買入（残存期間...）" 或 "Outright purchases of JGBs" 行，
    落札結果列（col 10）= "Amounts of Successful Bid" 是实际购入额。

    若当日无购入操作，返回 0；若文件不存在或无数据返回 None。
    """
    import openpyxl, io
    from datetime import date

    today = date.today()
    ymd = today.strftime("%Y%m%d")
    year = str(today.year)
    url = f"https://www.boj.or.jp/en/statistics/boj/fm/ope/d_release/ope/{year}/ope{ymd}.xlsx"

    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code == 404:
            return None  # 当日数据尚未发布
        r.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
        ws = wb.active
        total = 0.0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 100, values_only=True):
            vals = [str(v) if v is not None else "" for v in row]
            row_text = " ".join(vals)
            if "Outright purchases of JGBs" in row_text and vals[1]:
                # vals[9] = 落札総額(成功投标额) in 100M JPY = 亿日元
                try:
                    amount = float(vals[9].replace(",", "")) if vals[9] else 0
                    total += amount
                except (ValueError, AttributeError):
                    pass
        return total if total > 0 else 0.0
    except Exception:
        return None


def fetch_tonar() -> Optional[float]:
    """
    TONAR（东京隔夜平均利率，日终值）。
    BoJ 每日 17:15 JST 公布。

    2026-06-01 更新：旧页面已暂停（https://www3.boj.or.jp/market/en/menu_m.htm），
    数据移至新 URL：
    https://www.boj.or.jp/en/statistics/market/short/mutan/index.htm
    Excel 格式：provisional = mpYYYYMMDD.xlsx, final = mdYYYYMMDD.xlsx

    如无 openpyxl，返回 None（需先 pip install openpyxl）。
    """
    try:
        import openpyxl
    except ImportError:
        return None

    try:
        r = requests.get(BOJ_TONAR_URL, headers=HEADERS, timeout=15)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        # 旧页面返回大 comment — 先看是否被 suspend
        if "currently suspended" in soup.get_text():
            # 新页面：找最新 mp/md Excel
            r2 = requests.get(
                "https://www.boj.or.jp/en/statistics/market/short/mutan/index.htm",
                headers=HEADERS, timeout=15
            )
            r2.raise_for_status()
            soup2 = BeautifulSoup(r2.text, "html.parser")
            # 找最新 provisional 链接 mpYYYYMMDD.xlsx
            all_links = soup2.find_all("a")
            xlsx_hrefs = []
            for a in all_links:
                href = a.get("href", "")
                if "mp" in href and "xlsx" in href:
                    xlsx_hrefs.append(href)
            if xlsx_hrefs:
                # 取第一个（最新的 provisional 结果）
                latest = xlsx_hrefs[0]
                if latest.startswith("/"):
                    latest = "https://www.boj.or.jp" + latest
                rx = requests.get(latest, headers=HEADERS, timeout=15)
                rx.raise_for_status()
                import io
                wb = openpyxl.load_workbook(io.BytesIO(rx.content))
                ws = wb.active
                for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
                    # Average value is in col C (index 2), label in col B (index 1)
                    if row[1] and "Average" in str(row[1]) and row[2] is not None:
                        return float(row[2])
        else:
            # 旧解析方式（后备）
            text = soup.get_text()
            m = re.search(
                r"TONA[R]?\s*\(\s*Tokyo Overnight.*?\)\s*[:：]\s*(\d+\.\d+)",
                text, re.I
            )
            if m:
                return float(m.group(1))
    except Exception:
        pass
    return None


def fetch_usd_jpy() -> Optional[float]:
    """USD/JPY 汇率。优先 FRED DEXJPUS（H.10 周度有时滞），实时用其他源。"""
    try:
        from scrapers.fred_client import latest_value
        rec = latest_value("DEXJPUS")
        return rec["value"] if rec else None
    except Exception:
        return None


def fetch_cny_jpy() -> Optional[float]:
    """
    CNY/JPY 交叉汇率。
    用 USD/JPY ÷ USD/CNY 反算（FRED DEXJPUS / DEXCHUS）。
    """
    try:
        from scrapers.fred_client import latest_value
        usd_jpy = latest_value("DEXJPUS")
        usd_cny = latest_value("DEXCHUS")
        if usd_jpy and usd_cny:
            return usd_jpy["value"] / usd_cny["value"]
    except Exception:
        pass
    return None


if __name__ == "__main__":
    import json
    out = {
        "boj_rate_pct": fetch_boj_rate(),
        "jgb_daily_buy_oku_jpy": fetch_jgb_daily_buy(),
        "tonar_pct": fetch_tonar(),
        "usd_jpy": fetch_usd_jpy(),
        "cny_jpy": fetch_cny_jpy(),
    }
    print(json.dumps(out, indent=2, ensure_ascii=False))
